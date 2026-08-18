# -*- coding: utf-8 -*-
"""
auditoria_maquila.py — Auditoría de semielaborado de un almacén de maquila.
Proyecto RAGNAR · Deportivos Quini · 1 docena = 24 piezas

Balance:  Inv.Inicial + Traspasado + AjusteNeto − Entregado − Devuelto = Teórico
          Teórico vs Inventario físico reportado por el maquilero

Reglas aprendidas (auditoría Edgar Munguía, jul-2026):
 - Artículos CON 'Ensamble' (entrada) = packs terminados; su Traspaso(salida)
   es envío de packs a Quini, NO devolución de semielaborado.
 - Semielaborado puro = artículos SIN entradas por Ensamble.
 - Ajustes entrada/salida suelen ser sustituciones de códigos (efecto ~neto cero).
 - Docenas entregadas e inventario del maquilero: multiplicar ×24 para comparar.

ARCHIVOS DE ENTRADA (CONFIG):
  1. KARDEX:     'Kardex de los artículos' del almacén de la maquila (Microsip)
  2. RESUMEN:    RESUMEN_GENERAL_MAQUILAS.xlsx, hoja 'RESUMEN GENERAL ENSAMBLES'
  3. EXISTENCIA: inventario físico reportado (hoja 'INV': CODIGO, DOCENAS)

USO:  python auditoria_maquila.py
"""
import pandas as pd
import openpyxl
from quini_lib import (PZAS_POR_DOCENA, cargar_kardex_articulos, pivot_articulos,
                       extraer_codigo, escribir_encabezado, escribir_fila)

# ============ CONFIG ============
NOMBRE_MAQUILA     = "EDGAR MUNGUIA"      # como aparece en columna MAQUILA del resumen
ARCHIVO_KARDEX     = "Kardex_articulos_maquila.xlsx"
ARCHIVO_RESUMEN    = "RESUMEN_GENERAL_MAQUILAS.xlsx"
HOJA_RESUMEN       = "RESUMEN GENERAL ENSAMBLES"
ARCHIVO_EXISTENCIA = "existencia_maquila.xlsx"
HOJA_EXISTENCIA    = "INV"
FECHA_DESDE        = "2026-01-01"          # periodo del kardex
INV_INICIAL_DOC    = 0.0                   # inventario físico de arranque no cargado en Microsip
LINEAS_SEMI = ['Línea: CODIGOS','Línea: CALCETA DEPORTIVA','Línea: CALCETIN BEBES',
               'Línea: CALCETIN CABALLERO','Línea: CALCETIN DAMA','Línea: CALCETIN JR.',
               'Línea: CALCETIN NONES','Línea: TINES DEPORTIVOS']
ARCHIVO_SALIDA     = "Auditoria_Semielaborado_Maquila.xlsx"
# ================================

print(f"Auditoría: {NOMBRE_MAQUILA}")
df_movs, df_arts = cargar_kardex_articulos(ARCHIVO_KARDEX)
piv = pivot_articulos(df_movs, df_arts, LINEAS_SEMI)

semi  = piv[piv['Ens'] == 0]     # semielaborado puro
packs = piv[piv['Ens'] > 0]      # producto terminado ensamblado en el almacén

TE  = semi['TE'].sum()/PZAS_POR_DOCENA
UC  = semi['UC'].sum()/PZAS_POR_DOCENA
TS  = semi['TS'].sum()/PZAS_POR_DOCENA      # devoluciones reales de semielaborado
AjE = semi['AjE'].sum()/PZAS_POR_DOCENA
AjS = semi['AjS'].sum()/PZAS_POR_DOCENA
InvMicro = semi['Inv_Final'].sum()/PZAS_POR_DOCENA

# Entregas del resumen
df_res = pd.read_excel(ARCHIVO_RESUMEN, sheet_name=HOJA_RESUMEN, engine='calamine', header=0)
df_res.columns = ['MAQUILA','FECHA','NOTA','PEDIDO_OC','CODIGO','DESCRIPCION',
                  'DOC_ENTREGADAS','OBSERVACIONES'][:len(df_res.columns)]
df_res['DOC_ENTREGADAS'] = pd.to_numeric(df_res['DOC_ENTREGADAS'], errors='coerce').fillna(0)
df_res['FECHA'] = pd.to_datetime(df_res['FECHA'], errors='coerce')
mask = df_res['MAQUILA'].astype(str).str.upper().str.contains(NOMBRE_MAQUILA.split()[0].upper(), na=False)
df_e = df_res[mask & (df_res['FECHA'] >= FECHA_DESDE)]
ENTREGADO = df_e['DOC_ENTREGADAS'].sum()

# Inventario físico reportado
df_inv = pd.read_excel(ARCHIVO_EXISTENCIA, sheet_name=HOJA_EXISTENCIA, engine='calamine', header=0)
df_inv = df_inv.iloc[:, :2]
df_inv.columns = ['CODIGO','DOC_INV']
df_inv['DOC_INV'] = pd.to_numeric(df_inv['DOC_INV'], errors='coerce')
df_inv = df_inv.dropna(subset=['DOC_INV'])
df_inv = df_inv[df_inv['CODIGO'].notna()]
df_inv['CODIGO'] = df_inv['CODIGO'].astype(str).str.strip()
df_inv = df_inv.groupby('CODIGO')['DOC_INV'].sum().reset_index()
INV_REPORTADO = df_inv['DOC_INV'].sum()

DISPONIBLE = INV_INICIAL_DOC + TE + (AjE - AjS)
TEORICO = DISPONIBLE - ENTREGADO - TS
DIF = TEORICO - INV_REPORTADO

print(f"""
=== BALANCE SEMIELABORADO (docenas) ===
(+) Inventario inicial:        {INV_INICIAL_DOC:>12,.2f}
(+) Traspasado por Quini:      {TE:>12,.2f}
(+) Ajuste neto sustituciones: {AjE-AjS:>12,.2f}
    DISPONIBLE:                {DISPONIBLE:>12,.2f}
(–) Entregado ({FECHA_DESDE}→):{ENTREGADO:>12,.2f}
(–) Devuelto (semielaborado):  {TS:>12,.2f}
    TEÓRICO EN SU PODER:       {TEORICO:>12,.2f}
    Reportado físico:          {INV_REPORTADO:>12,.2f}
    DIFERENCIA:                {DIF:>12,.2f}  ({DIF*PZAS_POR_DOCENA:,.0f} pzas)
Referencias: Consumido Microsip (UC): {UC:,.2f} doc | Inv final Microsip: {InvMicro:,.2f} doc
Packs producidos: {packs['Ens'].sum()/PZAS_POR_DOCENA:,.2f} doc | Packs enviados: {packs['TS'].sum()/PZAS_POR_DOCENA:,.2f} doc
""")

# Comparativo de inventario por código
micro_cod = semi.groupby('Codigo')['Inv_Final'].sum().reset_index()
micro_cod.columns = ['CODIGO','Pzas_Micro']
micro_cod['Doc_Micro'] = micro_cod['Pzas_Micro']/PZAS_POR_DOCENA
comp = df_inv.merge(micro_cod, on='CODIGO', how='outer')
for c in ['DOC_INV','Doc_Micro','Pzas_Micro']: comp[c] = comp[c].fillna(0)
comp['Dif_Doc'] = comp['DOC_INV'] - comp['Doc_Micro']
comp = comp[(comp['DOC_INV']!=0)|(comp['Doc_Micro']!=0)].sort_values('Dif_Doc')

def sem_inv(d):
    if abs(d) <= 1: return ('Coincide','verde')
    if d < -1: return ('Microsip tiene MÁS (posibles ensambles sin capturar)','rojo')
    return ('Maquila reporta MÁS que Microsip','naranja')
comp[['Sem','Color']] = comp['Dif_Doc'].apply(lambda d: pd.Series(sem_inv(d)))

# ---------- Excel ----------
wb = openpyxl.Workbook()
from openpyxl.styles import Font as F
ws1 = wb.active; ws1.title = "Balance"
ws1.column_dimensions['A'].width=42; ws1.column_dimensions['B'].width=15; ws1.column_dimensions['C'].width=15
ws1.cell(1,1,f"AUDITORÍA SEMIELABORADO — {NOMBRE_MAQUILA}").font = F(bold=True,size=13,color="1F4E79")
filas = [("(+) Inventario inicial",INV_INICIAL_DOC),("(+) Traspasado por Quini",TE),
         ("(+) Ajuste neto (sustituciones)",AjE-AjS),("= DISPONIBLE",DISPONIBLE),
         ("(–) Entregado",ENTREGADO),("(–) Devuelto semielaborado",TS),
         ("= TEÓRICO EN SU PODER",TEORICO),("Reportado físico",INV_REPORTADO),
         ("DIFERENCIA",DIF),("",""),
         ("Ref: Consumido Microsip (UC)",UC),("Ref: Inv final Microsip",InvMicro),
         ("Ref: Packs producidos",packs['Ens'].sum()/PZAS_POR_DOCENA),
         ("Ref: Packs enviados",packs['TS'].sum()/PZAS_POR_DOCENA)]
for r,(a,b) in enumerate(filas,3):
    ws1.cell(r,1,a)
    if b != "":
        ws1.cell(r,2,round(b,2))
        ws1.cell(r,3,round(b*PZAS_POR_DOCENA,0)).font = F(color="595959", size=9)
    if str(a).startswith('=') or a=='DIFERENCIA':
        ws1.cell(r,1).font = F(bold=True); ws1.cell(r,2).font = F(bold=True)
ws1.cell(2,2,"DOCENAS").font=F(bold=True); ws1.cell(2,3,"PIEZAS").font=F(bold=True)

ws2 = wb.create_sheet("Inventario_por_Codigo")
escribir_encabezado(ws2, ['CODIGO','DOC MAQUILA','DOC MICROSIP','DIF DOC','DIF PZAS','SEMAFORO'],
                    [14,14,14,12,12,48])
for i,row in enumerate(comp.itertuples(),2):
    escribir_fila(ws2, i, [row.CODIGO, round(row.DOC_INV,2), round(row.Doc_Micro,2),
                           round(row.Dif_Doc,2), int(row.Dif_Doc*PZAS_POR_DOCENA), row.Sem], row.Color)
ws2.auto_filter.ref = f"A1:F{len(comp)+1}"

ws3 = wb.create_sheet("Detalle_Articulos")
escribir_encabezado(ws3, ['CODIGO','ARTICULO','LINEA','TIPO','TRASPASADO','CONSUMIDO','TS SALIDA',
                          'AJ ENT','AJ SAL','INV FINAL PZAS','INV FINAL DOC'], [12,50,22,14,12,12,12,10,10,13,12])
piv_o = piv.sort_values(['Codigo','Articulo'])
for i,row in enumerate(piv_o.itertuples(),2):
    tipo = 'PACK' if row.Ens>0 else 'SEMIELAB'
    escribir_fila(ws3, i, [row.Codigo, row.Articulo, row.Linea.replace('Línea: ',''), tipo,
        int(row.TE), int(row.UC), int(row.TS), int(row.AjE), int(row.AjS),
        int(row.Inv_Final), round(row.Inv_Final/PZAS_POR_DOCENA,2)],
        'amarillo' if row.Inv_Final>0 else 'gris')
ws3.auto_filter.ref = f"A1:K{len(piv_o)+1}"

wb.save(ARCHIVO_SALIDA)
print(f"Guardado: {ARCHIVO_SALIDA}")
