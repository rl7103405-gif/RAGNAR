# -*- coding: utf-8 -*-
"""
valuacion_inventarios.py — Consolida inventarios de cierre de las maquilas
y los valúa con costos (semielaborado y avíos), con descripción por código.
Proyecto RAGNAR · Deportivos Quini · 1 docena = 24 piezas

ARCHIVOS DE ENTRADA (CONFIG):
  1. INVENTARIOS: un archivo por maquila. Cada uno con columnas: CODIGO, DOCENAS
     (si alguna maquila reporta en piezas, márcala en UNIDAD_POR_MAQUILA)
  2. CATALOGO_COSTOS: export de Microsip con columnas: CODIGO, DESCRIPCION, COSTO
     - COSTO_UNIDAD indica si el costo del catálogo es por PIEZA o por DOCENA

SALIDA: Valuacion_Inventarios_Maquilas.xlsx
  - Resumen por maquila (docenas, piezas, valor $)
  - Detalle valuado por maquila y código
  - Códigos sin costo (para completar el catálogo)

USO:  python valuacion_inventarios.py
"""
import pandas as pd
import openpyxl
from quini_lib import PZAS_POR_DOCENA, escribir_encabezado, escribir_fila

# ============ CONFIG ============
INVENTARIOS = {
    # "NOMBRE MAQUILA": ("archivo.xlsx", "hoja o None"),
    "EDGAR MUNGUIA":  ("existencia_edgar_munguia.xlsx", "INV"),
    # "JAVIER MENDOZA": ("existencia_mendoza.xlsx", None),
    # "HUGO":           ("existencia_hugo.xlsx", None),
}
UNIDAD_POR_MAQUILA = {}          # ej. {"HUGO": "PIEZAS"} si alguien reporta piezas; default DOCENAS
CATALOGO_COSTOS = "catalogo_costos.xlsx"   # CODIGO, DESCRIPCION, COSTO
COSTO_UNIDAD    = "PIEZA"                   # "PIEZA" o "DOCENA"
ARCHIVO_SALIDA  = "Valuacion_Inventarios_Maquilas.xlsx"
# ================================

# --- Cargar catálogo de costos ---
cat = pd.read_excel(CATALOGO_COSTOS, engine='calamine', header=0)
cat = cat.iloc[:, :3]
cat.columns = ['CODIGO','DESCRIPCION','COSTO']
cat['CODIGO'] = cat['CODIGO'].astype(str).str.strip()
cat['COSTO'] = pd.to_numeric(cat['COSTO'], errors='coerce')
cat = cat.dropna(subset=['COSTO']).drop_duplicates('CODIGO')
print(f"Catálogo: {len(cat)} códigos con costo (por {COSTO_UNIDAD})")

# --- Cargar inventarios ---
frames = []
for maquila, (archivo, hoja) in INVENTARIOS.items():
    df = pd.read_excel(archivo, sheet_name=hoja or 0, engine='calamine', header=0)
    df = df.iloc[:, :2]
    df.columns = ['CODIGO','CANTIDAD']
    df['CODIGO'] = df['CODIGO'].astype(str).str.strip()
    df['CANTIDAD'] = pd.to_numeric(df['CANTIDAD'], errors='coerce')
    df = df.dropna(subset=['CANTIDAD'])
    df = df[df['CODIGO'].notna() & (df['CODIGO']!='nan')]
    unidad = UNIDAD_POR_MAQUILA.get(maquila, 'DOCENAS').upper()
    if unidad.startswith('PIEZA'):
        df['DOCENAS'] = df['CANTIDAD'] / PZAS_POR_DOCENA
    else:
        df['DOCENAS'] = df['CANTIDAD']
    df['PIEZAS'] = df['DOCENAS'] * PZAS_POR_DOCENA
    df['MAQUILA'] = maquila
    frames.append(df[['MAQUILA','CODIGO','DOCENAS','PIEZAS']])
    print(f"  {maquila}: {len(df)} códigos, {df['DOCENAS'].sum():,.2f} doc")

inv = pd.concat(frames, ignore_index=True)
inv = inv.groupby(['MAQUILA','CODIGO']).sum().reset_index()

# --- Valuar ---
inv = inv.merge(cat, on='CODIGO', how='left')
if COSTO_UNIDAD.upper().startswith('PIEZA'):
    inv['VALOR'] = inv['PIEZAS'] * inv['COSTO']
else:
    inv['VALOR'] = inv['DOCENAS'] * inv['COSTO']
inv['CON_COSTO'] = inv['COSTO'].notna()

sin_costo = inv[~inv['CON_COSTO']]
print(f"\nCódigos SIN costo en catálogo: {sin_costo['CODIGO'].nunique()} "
      f"({sin_costo['DOCENAS'].sum():,.2f} doc sin valuar)")

resumen = inv.groupby('MAQUILA').agg(
    Codigos=('CODIGO','nunique'), Docenas=('DOCENAS','sum'),
    Piezas=('PIEZAS','sum'), Valor=('VALOR','sum')).reset_index().sort_values('Valor', ascending=False)
print("\n=== RESUMEN ===")
print(resumen.to_string(index=False))
print(f"\nVALOR TOTAL EN MAQUILAS: ${inv['VALOR'].sum():,.2f}")

# ---------- Excel ----------
wb = openpyxl.Workbook()
from openpyxl.styles import Font as F
ws1 = wb.active; ws1.title = "Resumen"
ws1.cell(1,1,"VALUACIÓN DE INVENTARIOS EN MAQUILAS — CIERRE").font = F(bold=True,size=13,color="1F4E79")
escribir_encabezado(ws1, ['MAQUILA','CODIGOS','DOCENAS','PIEZAS','VALOR $'], [20,10,14,14,16], row=3)
r = 4
for row in resumen.itertuples():
    escribir_fila(ws1, r, [row.MAQUILA, row.Codigos, round(row.Docenas,2),
                           int(row.Piezas), round(row.Valor,2)], 'gris'); r += 1
escribir_fila(ws1, r, ['TOTAL', int(resumen['Codigos'].sum()), round(resumen['Docenas'].sum(),2),
                       int(resumen['Piezas'].sum()), round(resumen['Valor'].sum(),2)], 'verde')

ws2 = wb.create_sheet("Detalle_Valuado")
escribir_encabezado(ws2, ['MAQUILA','CODIGO','DESCRIPCION','DOCENAS','PIEZAS','COSTO UNIT','VALOR $','CON COSTO'],
                    [18,14,38,12,12,12,14,10])
for i,row in enumerate(inv.sort_values(['MAQUILA','VALOR'], ascending=[True,False]).itertuples(),2):
    escribir_fila(ws2, i, [row.MAQUILA, row.CODIGO,
        row.DESCRIPCION if pd.notna(row.DESCRIPCION) else 'SIN DESCRIPCIÓN',
        round(row.DOCENAS,2), int(row.PIEZAS),
        round(row.COSTO,4) if pd.notna(row.COSTO) else '',
        round(row.VALOR,2) if pd.notna(row.VALOR) else '',
        'SÍ' if row.CON_COSTO else 'NO'],
        'gris' if row.CON_COSTO else 'amarillo')
ws2.auto_filter.ref = f"A1:H{len(inv)+1}"

if len(sin_costo):
    ws3 = wb.create_sheet("Codigos_Sin_Costo")
    escribir_encabezado(ws3, ['CODIGO','MAQUILA(S)','DOCENAS TOTALES'], [14,30,16])
    sc = sin_costo.groupby('CODIGO').agg(
        Maquilas=('MAQUILA', lambda x: ', '.join(sorted(x.unique()))),
        Docenas=('DOCENAS','sum')).reset_index().sort_values('Docenas', ascending=False)
    for i,row in enumerate(sc.itertuples(),2):
        escribir_fila(ws3, i, [row.CODIGO, row.Maquilas, round(row.Docenas,2)], 'amarillo')

wb.save(ARCHIVO_SALIDA)
print(f"\nGuardado: {ARCHIVO_SALIDA}")
