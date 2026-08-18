# -*- coding: utf-8 -*-
"""
compulsa_triple.py — Auditoría física: SCQ vs REMISIONES vs MICROSIP (lotes)
Proyecto RAGNAR · Deportivos Quini

ARCHIVOS DE ENTRADA (edita las rutas en CONFIG):
  1. SCQ:        reporte de folios .xls (Seguimiento de folios, encabezado en fila 2)
  2. REMISIONES: BASE_REMISIONES.xlsx, hoja 'REMISIONES'
  3. MICROSIP:   COMPULSA_LOTES.xlsx (reporte 'Lotes con actividad' de procesos iniciales)

SALIDA: Auditoria_Fisica_Compulsa_Triple.xlsx
  - Resumen ejecutivo con semáforo
  - Detalle de todos los folios con remisión
  - Hoja de urgentes (vivo en 3 sistemas)
  - Una hoja por maquilero con discrepancias + columnas de firma

USO:  python compulsa_triple.py
"""
import pandas as pd
import openpyxl
from quini_lib import (cargar_scq, cargar_remisiones, agrupar_remisiones,
                       cargar_kardex_lotes, agregar_kardex_por_lote,
                       escribir_encabezado, escribir_fila)

# ============ CONFIG ============
ARCHIVO_SCQ        = "FOLIOS_SCQ.xls"
ARCHIVO_REMISIONES = "BASE_REMISIONES.xlsx"
ARCHIVO_MICROSIP   = "COMPULSA_LOTES.xlsx"
ARCHIVO_SALIDA     = "Auditoria_Fisica_Compulsa_Triple.xlsx"
TOLERANCIA_DOC     = 0.5
# ================================

print("Cargando fuentes...")
df_scq = cargar_scq(ARCHIVO_SCQ)
df_scq_u = df_scq.drop_duplicates('FOLIO')
df_rem = cargar_remisiones(ARCHIVO_REMISIONES)
df_rem_agg = agrupar_remisiones(df_rem)
dfk = cargar_kardex_lotes(ARCHIVO_MICROSIP)
df_micro = agregar_kardex_por_lote(dfk)

print(f"  SCQ: {len(df_scq_u)} folios | REM: {df_rem['FOLIO'].nunique()} folios | MICRO: {df_micro['FOLIO'].nunique()} lotes")

# Master: base = folios con remisión
m = df_rem_agg.merge(df_scq_u[['FOLIO','Docenas_Total','SCQ_Cerrado_OK','es_vivo','Estatus_SCQ',
                                'Codigo','Descripcion','Modelo','Color']], on='FOLIO', how='left')
m = m.merge(df_micro[['FOLIO','Doc_EP','Doc_TS','Doc_TE','Doc_Exist_Final']], on='FOLIO', how='left')
for c in ['Docenas_Total','Doc_REM','Doc_EP','Doc_TS','Doc_TE','Doc_Exist_Final']:
    m[c] = pd.to_numeric(m.get(c), errors='coerce').fillna(0)
m['En_SCQ'] = m['Estatus_SCQ'].notna()
m['En_MICRO'] = m['Doc_EP'] > 0
m['Dif_REM_vs_MICRO_TS'] = m['Doc_REM'] - m['Doc_TS']
m['Dif_REM_vs_SCQ'] = m['Doc_REM'] - m['Docenas_Total']

def semaforo(r):
    vivo = bool(r['es_vivo']) if pd.notna(r['es_vivo']) else False
    ok   = bool(r['SCQ_Cerrado_OK']) if pd.notna(r['SCQ_Cerrado_OK']) else False
    exist = r['Doc_Exist_Final']; d = r['Dif_REM_vs_MICRO_TS']
    if r['En_SCQ'] and vivo and r['En_MICRO'] and exist>0: return ('DISCREPANCIA: Vivo SCQ + Exist Microsip + REM enviada','rojo')
    if r['En_SCQ'] and vivo and r['En_MICRO']:             return ('SCQ vivo, Microsip cerrado','naranja')
    if r['En_SCQ'] and vivo:                                return ('SCQ vivo, sin movimiento Microsip','naranja')
    if r['En_SCQ'] and ok:
        if r['En_MICRO'] and exist==0 and abs(d)<=TOLERANCIA_DOC: return ('OK: SCQ cerrado + Microsip alineado','verde')
        if r['En_MICRO'] and exist>0:  return ('SCQ cerrado, Microsip con existencia','amarillo')
        if r['En_MICRO']:              return ('SCQ cerrado, diferencia REM vs Microsip','amarillo')
        return ('OK: SCQ cerrado','verde')
    if r['En_MICRO'] and exist>0:      return ('Sin SCQ, Microsip con existencia','amarillo')
    if r['En_MICRO']:
        if abs(d)<=TOLERANCIA_DOC:     return ('OK: Proceso correcto (Microsip alineado)','verde')
        return ('Sin SCQ, diferencia REM vs Microsip','amarillo')
    return ('OK: Proceso correcto (sin residuo)','verde')

m[['Semaforo','ColorSem']] = m.apply(lambda r: pd.Series(semaforo(r)), axis=1)
m = m.sort_values(['Maquilero','ColorSem','FOLIO'])

print("\nResultado:")
print(m['Semaforo'].value_counts().to_string())

# ---------- Excel ----------
wb = openpyxl.Workbook()
ws1 = wb.active; ws1.title = "Resumen"
from openpyxl.styles import Font as F
ws1.cell(1,1,"COMPULSA TRIPLE: SCQ · REMISIONES · MICROSIP").font = F(bold=True,size=13,color="1F4E79")
ws1.column_dimensions['A'].width=55; ws1.column_dimensions['B'].width=12
r = 3
for sem, cnt in m['Semaforo'].value_counts().items():
    ws1.cell(r,1,sem); ws1.cell(r,2,int(cnt)); r += 1

H = ['FOLIO','MAQUILERO','REMISION(ES)','OT','DESCRIPCION','DOC REM','DOC SCQ','ESTATUS SCQ',
     'DOC EP MICRO','DOC TS MICRO','DOC EXIST MICRO','DIF REM-MICRO','SEMAFORO']
W = [10,12,18,8,28,10,10,20,12,12,14,12,44]
def fila(row):
    return [row.FOLIO, row.Maquilero, row.Remisiones, row.OT,
            row.Descripcion_REM if pd.notna(row.Descripcion_REM) else '',
            round(row.Doc_REM,2), round(row.Docenas_Total,2),
            row.Estatus_SCQ if pd.notna(row.Estatus_SCQ) else 'No en SCQ',
            round(row.Doc_EP,2), round(row.Doc_TS,2), round(row.Doc_Exist_Final,2),
            round(row.Dif_REM_vs_MICRO_TS,2), row.Semaforo]

ws2 = wb.create_sheet("Detalle")
escribir_encabezado(ws2, H, W)
for i,row in enumerate(m.itertuples(),2): escribir_fila(ws2, i, fila(row), row.ColorSem)
ws2.auto_filter.ref = f"A1:M{len(m)+1}"

ws3 = wb.create_sheet("URGENTE_rojos")
mr = m[m['ColorSem']=='rojo']
escribir_encabezado(ws3, H+['FIRMA','OBSERVACIONES'], W+[20,30])
for i,row in enumerate(mr.itertuples(),2): escribir_fila(ws3, i, fila(row)+['',''], 'rojo')

md = m[m['ColorSem']!='verde']
for maq in sorted(md['Maquilero'].dropna().unique()):
    dfm = md[md['Maquilero']==maq]
    ws = wb.create_sheet(f"Disc_{str(maq)[:24]}")
    escribir_encabezado(ws, H+['FIRMA','OBSERVACIONES'], W+[20,30])
    for i,row in enumerate(dfm.itertuples(),2): escribir_fila(ws, i, fila(row)+['',''], row.ColorSem)
    ws.auto_filter.ref = f"A1:O{len(dfm)+1}"

wb.save(ARCHIVO_SALIDA)
print(f"\nGuardado: {ARCHIVO_SALIDA}")
