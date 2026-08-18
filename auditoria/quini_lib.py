# -*- coding: utf-8 -*-
"""
quini_lib.py — Funciones compartidas para auditorías Deportivos Quini
Proyecto RAGNAR · Requiere: pandas, openpyxl, xlrd, python-calamine
    pip install pandas openpyxl xlrd python-calamine --break-system-packages
Regla de unidades: 1 docena = 12 pares = 24 piezas
"""
import pandas as pd
import re

PZAS_POR_DOCENA = 24

# ---------------------------------------------------------------- SCQ
def cargar_scq(path):
    """FOLIOS SCQ (.xls). Devuelve DataFrame con FOLIO, Docenas_Total,
    Cerrado/Volteado/Hormado/Pareado, es_vivo, Estatus_SCQ."""
    df = pd.read_excel(path, engine="xlrd", header=None, skiprows=2)
    df.columns = ['Fecha','FOLIO','Codigo','OT_Descripcion','Descripcion','Modelo',
                  'Color','Docenas_Total','Col8','Cerrado','Volteado','Hormado','Pareado']
    df['FOLIO'] = pd.to_numeric(df['FOLIO'], errors='coerce')
    df = df.dropna(subset=['FOLIO'])
    df['FOLIO'] = df['FOLIO'].astype(int)
    df['Docenas_Total'] = pd.to_numeric(df['Docenas_Total'], errors='coerce').fillna(0)
    df['SCQ_Cerrado_OK'] = (df['Cerrado']==1)&(df['Volteado']==1)&(df['Hormado']==1)
    df['es_vivo'] = ~df['SCQ_Cerrado_OK']
    df['Estatus_SCQ'] = df.apply(lambda r: f"C={'OK' if r['Cerrado']==1 else 'X'} "
        f"V={'OK' if r['Volteado']==1 else 'X'} H={'OK' if r['Hormado']==1 else 'X'}", axis=1)
    return df

# --------------------------------------------------------- REMISIONES
def cargar_remisiones(path, hoja='REMISIONES'):
    """BASE_REMISIONES.xlsx. Devuelve DataFrame línea x línea."""
    df = pd.read_excel(path, sheet_name=hoja, header=None, skiprows=1)
    df.columns = ['FOLIO','PESO','Codigo','Descripcion','Modelo','Talla','Color',
                  'Referencia','Linea','Docenas','Unidad','UPC','OT','REMISION','MAQUILERO','FECHA']
    df['FOLIO'] = pd.to_numeric(df['FOLIO'], errors='coerce')
    df = df.dropna(subset=['FOLIO'])
    df['FOLIO'] = df['FOLIO'].astype(int)
    df['REMISION'] = pd.to_numeric(df['REMISION'], errors='coerce')
    df['Docenas'] = pd.to_numeric(df['Docenas'], errors='coerce')
    return df

def agrupar_remisiones(df_rem):
    """Agrega remisiones por FOLIO."""
    return df_rem.groupby('FOLIO').agg(
        Remisiones=('REMISION', lambda x: ', '.join(str(int(v)) for v in sorted(x.dropna().unique()))),
        Maquilero=('MAQUILERO','first'),
        Doc_REM=('Docenas','sum'),
        OT=('OT','first'),
        Descripcion_REM=('Descripcion','first'),
    ).reset_index()

# ------------------------------------------- KARDEX DE LOTES (Microsip)
def cargar_kardex_lotes(path, skip_headers=None):
    """COMPULSA_LOTES.xlsx — reporte 'Lotes con actividad' de Microsip.
    Devuelve DataFrame de movimientos: LOTE, Existencia_Final, Fecha, Concepto, Unidades(pzas)."""
    if skip_headers is None:
        skip_headers = ['Deportivos Quini','Lotes','PROCESOS INICIALES','Lotes con actividad','Artículo']
    df = pd.read_excel(path, engine='calamine', header=None)
    records = []
    lote = None; exist = None
    for _, row in df.iterrows():
        v0, v1 = row[0], row[1]
        s0 = str(v0).strip(); s1 = str(v1).strip()
        if s0 in skip_headers or s0.startswith('Del ') or s1 in ['Lote','Fecha']:
            continue
        if pd.notna(v0) and s0 not in ['','nan']:
            lote = None; continue                      # encabezado de artículo
        if pd.isna(v0) and pd.notna(v1) and s1 not in ['','nan']:
            try:
                lote = int(float(s1))
                exist = row[7] if pd.notna(row[7]) else None
                if pd.notna(row[8]) and pd.notna(row[10]):
                    records.append({'LOTE':lote,'Existencia_Final':exist,'Fecha':row[8],
                                    'Concepto':str(row[10]).strip(),
                                    'Unidades':row[17] if pd.notna(row[17]) else 0})
            except Exception:
                pass
            continue
        if pd.isna(v0) and pd.isna(v1) and pd.notna(row[8]) and pd.notna(row[10]):
            records.append({'LOTE':lote,'Existencia_Final':exist,'Fecha':row[8],
                            'Concepto':str(row[10]).strip(),
                            'Unidades':row[17] if pd.notna(row[17]) else 0})
    dfk = pd.DataFrame(records)
    dfk['LOTE'] = pd.to_numeric(dfk['LOTE'], errors='coerce')
    dfk['Unidades'] = pd.to_numeric(dfk['Unidades'], errors='coerce').fillna(0)
    dfk['Existencia_Final'] = pd.to_numeric(dfk['Existencia_Final'], errors='coerce')
    return dfk

def clasificar_concepto(c):
    c = c.lower()
    if 'entrada por produccion' in c: return 'ep'
    if 'traspaso (salida)' in c: return 'ts'
    if 'traspaso (entrada)' in c: return 'te'
    if 'ajuste (salida)' in c or 'inv. físico (salida)' in c: return 'as'
    if 'ajuste (entrada)' in c or 'inv. físico (entrada)' in c: return 'ae'
    if 'uso componente' in c: return 'uc'
    if 'ensamble' in c: return 'ens'
    return 'otro'

def agregar_kardex_por_lote(dfk):
    """Pivot por LOTE con piezas por tipo de movimiento + docenas."""
    dfk = dfk.copy()
    dfk['tipo'] = dfk['Concepto'].apply(clasificar_concepto)
    agg = dfk.groupby(['LOTE','Existencia_Final']).apply(lambda g: pd.Series({
        'Pzas_EP': g[g['tipo']=='ep']['Unidades'].sum(),
        'Pzas_TS': g[g['tipo']=='ts']['Unidades'].sum(),
        'Pzas_TE': g[g['tipo']=='te']['Unidades'].sum(),
        'Pzas_AS': g[g['tipo']=='as']['Unidades'].sum(),
        'Pzas_AE': g[g['tipo']=='ae']['Unidades'].sum(),
    })).reset_index()
    for p in ['EP','TS','TE','AS','AE']:
        agg[f'Doc_{p}'] = agg[f'Pzas_{p}'] / PZAS_POR_DOCENA
    agg['Doc_Exist_Final'] = agg['Existencia_Final'] / PZAS_POR_DOCENA
    return agg.rename(columns={'LOTE':'FOLIO'})

# --------------------------------- KARDEX DE ARTÍCULOS (almacén maquila)
def cargar_kardex_articulos(path):
    """'Kardex de los artículos' de un almacén de maquila.
    Devuelve (df_movimientos, df_articulos)."""
    df = pd.read_excel(path, engine='calamine', header=None)
    linea = None; art = None
    arts = []; movs = []
    for _, row in df.iterrows():
        v0 = str(row[0]).strip() if pd.notna(row[0]) else None
        if v0 and v0.startswith('Línea:'):
            linea = v0; art = None; continue
        if v0 and pd.notna(row[6]) and str(row[6]).strip() in ['PIEZA','PAR','DOCENA','PZA']:
            art = v0
            arts.append({'Articulo':art,'Linea':linea,'Unidad':str(row[6]).strip(),
                'Inv_Inicial':pd.to_numeric(row[10],errors='coerce') or 0,
                'Inv_Final':pd.to_numeric(row[17],errors='coerce') or 0})
            continue
        if art and pd.isna(row[0]) and pd.notna(row[1]) and pd.notna(row[2]):
            ent = pd.to_numeric(row[12], errors='coerce'); sal = pd.to_numeric(row[15], errors='coerce')
            movs.append({'Articulo':art,'Linea':linea,
                'Fecha':pd.to_datetime(row[1],errors='coerce'),
                'Concepto':str(row[2]).strip(),
                'Entrada':ent if pd.notna(ent) else 0,'Salida':sal if pd.notna(sal) else 0})
    return pd.DataFrame(movs), pd.DataFrame(arts)

def extraer_codigo(nombre_articulo):
    """Extrae código tipo 2533-H / 4664 / 1118-I del inicio del nombre."""
    m = re.match(r'^(\d{3,4}(?:-[A-Z0-9])?)', str(nombre_articulo))
    return m.group(1) if m else str(nombre_articulo)[:15]

def pivot_articulos(df_movs, df_arts, lineas=None):
    """Pivot por artículo con TE/UC/TS/AjE/AjS/Ens + Inv_Final. lineas=None → todas."""
    dm = df_movs if lineas is None else df_movs[df_movs['Linea'].isin(lineas)]
    piv = dm.pivot_table(index=['Articulo','Linea'], columns='Concepto',
                         values=['Entrada','Salida'], aggfunc='sum', fill_value=0)
    piv.columns = [f"{a}|{b}" for a,b in piv.columns]
    piv = piv.reset_index()
    def C(n): return piv[n] if n in piv.columns else 0
    piv['TE']  = C('Entrada|Traspaso (entrada)')
    piv['Ens'] = C('Entrada|Ensamble')
    piv['UC']  = C('Salida|Uso componente')
    piv['TS']  = C('Salida|Traspaso (salida)')
    piv['AjE'] = C('Entrada|Ajuste (entrada)') + C('Entrada|Inv. físico (entrada)')
    piv['AjS'] = C('Salida|Ajuste (salida)') + C('Salida|Inv. físico (salida)')
    inv = df_arts.set_index('Articulo')['Inv_Final'].to_dict()
    piv['Inv_Final'] = piv['Articulo'].map(inv).fillna(0)
    piv['Codigo'] = piv['Articulo'].apply(extraer_codigo)
    return piv

# --------------------------------------------------------------- EXCEL
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)

COLORES = {'rojo':('FF0000','FFFFFF'),'naranja':('FF8C00','FFFFFF'),
           'amarillo':('FFD700','000000'),'verde':('C6EFCE','000000'),
           'gris':('F2F2F2','000000')}

def escribir_encabezado(ws, headers, widths, row=1):
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(row,c,h); cell.fill=HDR_FILL; cell.font=HDR_FONT; cell.alignment=CENTER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = f'A{row+1}'

def escribir_fila(ws, r, vals, color='gris'):
    bg, fc = COLORES.get(color, COLORES['gris'])
    fill = PatternFill("solid", fgColor=bg); fnt = Font(color=fc, size=9)
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v); cell.fill = fill; cell.font = fnt
