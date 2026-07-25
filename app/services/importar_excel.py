"""
Importacion de los dos Excel diarios que alimentan la estacion de Validacion:

1. 'Folios ruteo DD.MM.AA.xlsx' — folios esperados (ACUMULATIVO: upsert por folio).
2. 'Codigos_Productos_Quini DD.MM.AA.xlsx' — BOM del que se agrega el peso
   teorico unitario por codigo (SNAPSHOT: cada importacion reemplaza el
   catalogo completo, previa validacion de TODO el archivo).

Ambas funciones validan el archivo completo ANTES de escribir en la base:
una fila invalida rechaza la importacion entera con un reporte de errores,
para que el catalogo/los folios nunca queden a medias.
"""
import logging
import math
import re
from datetime import datetime
from io import BytesIO

import openpyxl

from app.models import Bulto, CatalogoPeso, FolioRuteo
from app.validation import canonizar_folio, PATRON_FOLIO, PATRON_PEDIDO

logger = logging.getLogger("importar_excel")

# Limites defensivos: un XLSX es un ZIP y puede descomprimir a mucho mas de
# lo que pesa el archivo (zip bomb). Ademas del tamano del upload (validado
# en el router), se limita el numero de filas procesables.
MAX_FILAS_FOLIOS = 50_000
MAX_FILAS_CATALOGO = 1_000_000
MAX_ERRORES_REPORTADOS = 20

ENCABEZADOS_FOLIOS = [
    "folio", "fecha", "codigo", "docenas", "pares", "total",
    "fecha captura", "fecha actualizacion", "descripcion", "pedido",
    "num. ruta", "nombre_guia",
]
# Del catalogo solo se usan la primera columna (CodigoProducto) y la ultima
# (Cantidad); se validan sus nombres para detectar un archivo equivocado.
COLUMNA_CODIGO_CATALOGO = "codigoproducto"
COLUMNA_CANTIDAD_CATALOGO = "cantidad"


class ErrorImportacion(Exception):
    """Archivo rechazado; .errores trae el detalle por fila."""

    def __init__(self, mensaje: str, errores: list[str] | None = None):
        super().__init__(mensaje)
        self.mensaje = mensaje
        self.errores = errores or []


def _abrir_hoja(contenido: bytes):
    try:
        libro = openpyxl.load_workbook(
            BytesIO(contenido), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        raise ErrorImportacion(f"No se pudo abrir el Excel: {exc}") from exc
    if not libro.sheetnames:
        raise ErrorImportacion("El Excel no tiene hojas")
    return libro, libro[libro.sheetnames[0]]


def _texto_celda(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def _numero_o_none(valor):
    if valor is None or valor == "":
        return None
    # bool hereda de int en Python: sin excluirlo, una celda con TRUE/FALSE
    # (o 0/1 que Excel guarda como booleano) se aceptaria como numero valido.
    if isinstance(valor, (int, float)) and not isinstance(valor, bool) and math.isfinite(float(valor)):
        return float(valor)
    return "invalido"


def _valor(fila, i):
    """Acceso seguro a una celda: filas mas cortas que lo esperado no truenan."""
    return fila[i] if i < len(fila) else None


def _sanear_nombre_archivo(nombre: str | None) -> str:
    """Recorta y quita caracteres de control antes de guardarlo como archivo_origen."""
    nombre = (nombre or "").strip()
    nombre = "".join(c for c in nombre if ord(c) >= 32 and ord(c) != 127)
    return nombre[:120]


_PATRON_FECHA_NOMBRE_ARCHIVO = re.compile(r"(\d{1,2})[._-](\d{1,2})[._-](\d{2,4})")


def _fecha_desde_nombre_archivo(nombre: str | None) -> datetime | None:
    """Extrae una fecha DD.MM.AA / DD_MM_AA del nombre del archivo, si trae una.

    Convencion usada en los Excel diarios (p.ej. 'Codigos_Productos_Quini_16_07_26.xlsx',
    'Folios ruteo 15.07.26.xlsx'). Devuelve None si el nombre no trae un patron
    reconocible o la fecha no es valida (no bloquea la importacion: solo se usa
    como señal auxiliar para no reemplazar un catalogo mas nuevo por error).
    """
    if not nombre:
        return None
    match = _PATRON_FECHA_NOMBRE_ARCHIVO.search(nombre)
    if not match:
        return None
    dia, mes, anio = (int(v) for v in match.groups())
    if anio < 100:
        anio += 2000
    try:
        return datetime(anio, mes, dia)
    except ValueError:
        return None


def parsear_folios_ruteo(contenido: bytes, nombre_archivo: str) -> dict:
    """Valida el Excel completo de folios de ruteo y arma la estructura a escribir.

    No toca la base de datos: se puede ejecutar fuera del candado de escritura.
    """
    nombre_archivo = _sanear_nombre_archivo(nombre_archivo)
    libro, hoja = _abrir_hoja(contenido)
    try:
        filas = hoja.iter_rows(values_only=True)
        encabezados = next(filas, None)
        if encabezados is None:
            raise ErrorImportacion("El Excel está vacío")
        normalizados = [_texto_celda(c).lower() for c in encabezados[: len(ENCABEZADOS_FOLIOS)]]
        if normalizados != ENCABEZADOS_FOLIOS:
            raise ErrorImportacion(
                "Los encabezados no coinciden con el formato esperado de 'Folios ruteo'. "
                f"Esperado: {ENCABEZADOS_FOLIOS}. Encontrado: {normalizados}"
            )

        registros = {}
        errores = []
        for num_fila, fila in enumerate(filas, start=2):
            if num_fila - 1 > MAX_FILAS_FOLIOS:
                raise ErrorImportacion(f"El archivo excede el máximo de {MAX_FILAS_FOLIOS} filas")
            if fila is None or all(c is None or _texto_celda(c) == "" for c in fila):
                continue  # fila totalmente vacía (colas típicas de Excel)

            folio_crudo = _texto_celda(_valor(fila, 0))
            codigo = _texto_celda(_valor(fila, 2)).upper()
            docenas = _numero_o_none(_valor(fila, 3))
            pares = _numero_o_none(_valor(fila, 4))
            total = _numero_o_none(_valor(fila, 5))
            fecha_actualizacion_celda = _valor(fila, 7)
            pedido = _texto_celda(_valor(fila, 9))
            nombre_guia = _texto_celda(_valor(fila, 11))

            fecha_actualizacion_invalida = (
                fecha_actualizacion_celda is not None
                and _texto_celda(fecha_actualizacion_celda) != ""
                and not isinstance(fecha_actualizacion_celda, datetime)
            )

            if not folio_crudo:
                errores.append(f"Fila {num_fila}: folio vacío")
            elif len(folio_crudo) > 50 or not PATRON_FOLIO.fullmatch(folio_crudo):
                errores.append(
                    f"Fila {num_fila}: folio inválido (usa sólo letras, números, punto, guion o guion bajo, "
                    "máximo 50 caracteres) — quedaría inaccesible en Validación/Embarque"
                )
            elif not codigo:
                errores.append(f"Fila {num_fila}: código vacío")
            elif docenas == "invalido" or pares == "invalido" or total == "invalido":
                errores.append(f"Fila {num_fila}: docenas/pares/total no numérico")
            elif total is not None and total < 0 or docenas is not None and docenas < 0:
                errores.append(f"Fila {num_fila}: cantidades negativas")
            elif fecha_actualizacion_invalida:
                errores.append(f"Fila {num_fila}: 'Fecha Actualizacion' no es una fecha válida")
            elif pedido and (len(pedido) > 100 or not PATRON_PEDIDO.fullmatch(pedido)):
                errores.append(
                    f"Fila {num_fila}: pedido inválido (usa sólo letras, números, punto, guion o guion bajo, "
                    "máximo 100 caracteres)"
                )
            elif (total or docenas or 0) <= 0:
                errores.append(f"Fila {num_fila}: el folio no tiene docenas ni total mayor a cero")
            else:
                folio = canonizar_folio(folio_crudo)
                if folio in registros:
                    errores.append(f"Fila {num_fila}: folio {folio} duplicado dentro del archivo")
                else:
                    total_final = total
                    if (total_final is None or total_final == 0) and docenas is not None and docenas > 0:
                        # Un total 0/ausente con docenas capturadas es inutilizable
                        # para el peso teórico; usa docenas como respaldo.
                        total_final = docenas
                    registros[folio] = {
                        "codigo": codigo,
                        "docenas": docenas,
                        "pares": pares,
                        "total": total_final,
                        "pedido_id": pedido or None,
                        "nombre_guia": nombre_guia or None,
                        "fecha": _valor(fila, 1) if isinstance(_valor(fila, 1), datetime) else None,
                        "fecha_actualizacion": fecha_actualizacion_celda if isinstance(fecha_actualizacion_celda, datetime) else None,
                    }
            if len(errores) >= MAX_ERRORES_REPORTADOS:
                break

        if errores:
            raise ErrorImportacion(
                "El archivo tiene filas inválidas y NO se importó nada", errores
            )
        if not registros:
            raise ErrorImportacion("El archivo no contiene folios")
    finally:
        libro.close()

    return {"nombre_archivo": nombre_archivo, "registros": registros}


def escribir_folios_ruteo(datos_parseados: dict, db) -> dict:
    """Escribe (upsert) los folios ya validados por parsear_folios_ruteo.

    Se ejecuta dentro del candado_escritura: no vuelve a leer ni validar el
    Excel, solo toca la base de datos.
    """
    nombre_archivo = datos_parseados["nombre_archivo"]
    registros = datos_parseados["registros"]

    ahora = datetime.now()
    insertados = actualizados = omitidos_viejos = omitidos_sin_fecha = 0
    conflictos = []
    for folio, datos in registros.items():
        existente = db.query(FolioRuteo).filter(FolioRuteo.folio == folio).first()
        if existente is None:
            # Si ya existe un bulto avanzado con este folio y sus datos difieren
            # de los que trae el Excel, se reporta el conflicto (informativo);
            # el folio se importa de todas formas.
            bulto = db.query(Bulto).filter(Bulto.folio == folio).first()
            if bulto is not None and bulto.estatus != "produccion":
                docenas_excel = datos["docenas"] if datos["docenas"] is not None else datos["total"]
                difiere = (
                    (bulto.codigo_producto or "") != (datos["codigo"] or "")
                    or (bulto.docenas or 0) != (docenas_excel or 0)
                    or (bulto.pedido_id or "") != (datos["pedido_id"] or "")
                )
                if difiere:
                    conflictos.append(
                        f"Folio {folio}: ya existe como bulto en estatus '{bulto.estatus}' con datos "
                        f"distintos a los del Excel (código {bulto.codigo_producto}->{datos['codigo']}, "
                        f"docenas {bulto.docenas}->{docenas_excel}, pedido {bulto.pedido_id}->{datos['pedido_id']}). "
                        "Se importó el folio de todas formas."
                    )
            db.add(FolioRuteo(
                folio=folio, archivo_origen=nombre_archivo, fecha_importacion=ahora, **datos,
            ))
            insertados += 1
            continue

        # No revertir datos con un Excel mas viejo que lo ya importado.
        if (
            datos["fecha_actualizacion"] is not None
            and existente.fecha_actualizacion is not None
            and datos["fecha_actualizacion"] < existente.fecha_actualizacion
        ):
            omitidos_viejos += 1
            continue

        # Si el Excel entrante no trae 'Fecha Actualizacion' no se sabe si es
        # mas nuevo o mas viejo que lo ya importado: no se sobrescribe a
        # ciegas (antes esto pisaba codigo/total/pedido conservando la fecha
        # vieja, ocultando que el dato se habia revertido).
        if datos["fecha_actualizacion"] is None and existente.fecha_actualizacion is not None:
            logger.warning(
                "Folio %s: el archivo '%s' no trae 'Fecha Actualizacion'; se omite la actualización "
                "para no revertir datos ya importados con fecha conocida.",
                folio, nombre_archivo,
            )
            omitidos_sin_fecha += 1
            continue

        # Si el folio ya fue validado fisicamente y el Excel ahora dice otra
        # cosa, no se cambia en silencio: se reporta el conflicto y el bulto
        # conserva el snapshot con el que fue validado.
        bulto = db.query(Bulto).filter(Bulto.folio == folio).first()
        cambia = (
            existente.codigo != datos["codigo"]
            or (existente.total or 0) != (datos["total"] or 0)
            or (existente.pedido_id or "") != (datos["pedido_id"] or "")
        )
        if bulto is not None and bulto.estatus != "produccion" and cambia:
            conflictos.append(
                f"Folio {folio}: ya fue procesado como '{bulto.estatus}' y el Excel trae datos "
                f"distintos (código {existente.codigo}->{datos['codigo']}, "
                f"total {existente.total}->{datos['total']}, pedido {existente.pedido_id}->{datos['pedido_id']}). "
                "No se modificó."
            )
            continue

        for campo, valor in datos.items():
            # No borrar una fecha_actualizacion ya conocida con un valor
            # ausente: solo se avanza hacia adelante, nunca a None.
            if campo == "fecha_actualizacion" and valor is None:
                continue
            setattr(existente, campo, valor)
        existente.archivo_origen = nombre_archivo
        existente.fecha_importacion = ahora
        actualizados += 1

    return {
        "archivo": nombre_archivo,
        "folios_en_archivo": len(registros),
        "insertados": insertados,
        "actualizados": actualizados,
        "omitidos_por_viejos": omitidos_viejos,
        "omitidos_por_falta_fecha": omitidos_sin_fecha,
        "conflictos": conflictos,
    }


def parsear_catalogo_pesos(contenido: bytes, nombre_archivo: str) -> dict:
    """Valida el BOM completo del catálogo de pesos. No toca la base de datos."""
    nombre_archivo = _sanear_nombre_archivo(nombre_archivo)
    libro, hoja = _abrir_hoja(contenido)
    try:
        filas = hoja.iter_rows(values_only=True)
        encabezados = next(filas, None)
        if encabezados is None:
            raise ErrorImportacion("El Excel está vacío")
        nombres = [_texto_celda(c).lower() for c in encabezados]
        if not nombres or nombres[0] != COLUMNA_CODIGO_CATALOGO or nombres[-1] != COLUMNA_CANTIDAD_CATALOGO:
            raise ErrorImportacion(
                "Los encabezados no coinciden con el formato esperado del catálogo "
                f"(primera columna '{COLUMNA_CODIGO_CATALOGO}', última '{COLUMNA_CANTIDAD_CATALOGO}'). "
                f"Encontrado: primera '{nombres[0] if nombres else ''}', última '{nombres[-1] if nombres else ''}'"
            )
        indice_cantidad = len(nombres) - 1

        suma = {}
        materiales = {}
        errores = []
        filas_leidas = 0
        for num_fila, fila in enumerate(filas, start=2):
            filas_leidas += 1
            if filas_leidas > MAX_FILAS_CATALOGO:
                raise ErrorImportacion(f"El archivo excede el máximo de {MAX_FILAS_CATALOGO} filas")
            if fila is None or all(c is None or _texto_celda(c) == "" for c in fila):
                continue

            codigo = _texto_celda(fila[0]).upper()
            cantidad = fila[indice_cantidad] if len(fila) > indice_cantidad else None
            if not codigo:
                errores.append(f"Fila {num_fila}: código vacío")
            elif (
                not isinstance(cantidad, (int, float))
                or isinstance(cantidad, bool)
                or not math.isfinite(float(cantidad))
            ):
                errores.append(f"Fila {num_fila}: cantidad no numérica para el código {codigo}")
            elif float(cantidad) < 0:
                errores.append(f"Fila {num_fila}: cantidad negativa para el código {codigo}")
            else:
                suma[codigo] = suma.get(codigo, 0.0) + float(cantidad)
                materiales[codigo] = materiales.get(codigo, 0) + 1
            if len(errores) >= MAX_ERRORES_REPORTADOS:
                break

        if errores:
            raise ErrorImportacion(
                "El catálogo tiene filas inválidas y NO se importó nada", errores
            )
        if not suma:
            raise ErrorImportacion("El catálogo no contiene códigos")
    finally:
        libro.close()

    # Un total <= 0 (p. ej. materiales que se cancelan entre si) es inutilizable
    # como peso teorico: se excluye del catalogo mas no rechaza el archivo (hay
    # codigos de muestra legitimos con cantidades muy pequeñas, no cero).
    codigos_sin_peso = sorted(codigo for codigo, total in suma.items() if total <= 0)
    suma_valida = {codigo: total for codigo, total in suma.items() if total > 0}
    materiales_valida = {codigo: materiales[codigo] for codigo in suma_valida}

    # Guardia contra un Excel patológico (p. ej. columnas movidas apuntando a
    # una columna de ceros): si NINGÚN código tiene peso utilizable, importar
    # vaciaría el catálogo entero y detendría toda la estación de Validación.
    if not suma_valida:
        raise ErrorImportacion(
            "Ningún código del archivo tiene peso mayor a cero: el catálogo NO se reemplazó. "
            "Revisa que el Excel sea el correcto y que la columna Cantidad tenga valores."
        )

    return {
        "nombre_archivo": nombre_archivo,
        "fecha_archivo": _fecha_desde_nombre_archivo(nombre_archivo),
        "filas_leidas": filas_leidas,
        "suma": suma_valida,
        "materiales": materiales_valida,
        "codigos_sin_peso": codigos_sin_peso,
    }


def escribir_catalogo_pesos(
    datos_parseados: dict, db, sobrescribir: bool = False, marca_confirmada: str | None = None
) -> dict:
    """Reemplaza el catálogo completo con lo ya validado por parsear_catalogo_pesos.

    Se ejecuta dentro del candado_escritura: reemplazo transaccional (borra e
    inserta en la misma transacción; si algo falla, el rollback conserva el
    catálogo anterior).
    """
    nombre_archivo = datos_parseados["nombre_archivo"]
    fecha_archivo_nuevo = datos_parseados.get("fecha_archivo")
    suma = datos_parseados["suma"]
    materiales = datos_parseados["materiales"]

    # El catalogo no trae una fecha por fila (es un BOM, no un Excel con
    # columna de fecha): la unica señal disponible es la fecha en el nombre
    # del archivo. Si el catalogo ya cargado parece mas nuevo que el archivo
    # que se va a importar, no se borra a ciegas (mismo patron de
    # marca_confirmada/sobrescribir que usa el modulo de Validacion).
    catalogo_existente = db.query(CatalogoPeso).first()
    if catalogo_existente is not None:
        marca_actual = catalogo_existente.archivo_origen
        fecha_archivo_actual = _fecha_desde_nombre_archivo(marca_actual)
        catalogo_mas_nuevo = (
            fecha_archivo_nuevo is not None
            and fecha_archivo_actual is not None
            and fecha_archivo_actual > fecha_archivo_nuevo
        )
        if catalogo_mas_nuevo and not sobrescribir:
            return {
                "ok": False,
                "requiere_sobrescribir": True,
                "marca": marca_actual,
                "mensaje": (
                    f"El catálogo cargado ('{marca_actual}') parece más reciente que el archivo que "
                    f"estás importando ('{nombre_archivo}'). ¿Reemplazarlo de todas formas?"
                ),
            }
        if catalogo_mas_nuevo and sobrescribir and marca_confirmada != marca_actual:
            # Entre el aviso y esta confirmacion, alguien mas reemplazo el
            # catalogo: no se pisa a ciegas, se vuelve a pedir confirmacion.
            return {
                "ok": False,
                "requiere_sobrescribir": True,
                "marca": marca_actual,
                "mensaje": (
                    f"El catálogo cambió mientras confirmabas (ahora es '{marca_actual}'). "
                    "Revisa y confirma de nuevo."
                ),
            }

    ahora = datetime.now()
    db.query(CatalogoPeso).delete()
    mapeos = [
        {
            "codigo": codigo,
            "peso_unitario_gramos": round(peso, 4),
            "num_materiales": materiales[codigo],
            "archivo_origen": nombre_archivo,
            "fecha_importacion": ahora,
        }
        for codigo, peso in suma.items()
    ]
    if mapeos:
        db.bulk_insert_mappings(CatalogoPeso, mapeos)

    return {
        "archivo": nombre_archivo,
        "filas_procesadas": datos_parseados["filas_leidas"],
        "codigos_importados": len(mapeos),
        "codigos_sin_peso": datos_parseados["codigos_sin_peso"],
    }
